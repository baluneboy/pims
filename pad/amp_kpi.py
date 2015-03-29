#!/usr/bin/env python

import re
import os
import sys
import csv
import numpy as np
import pandas as pd
import datetime
from cStringIO import StringIO
from pims.utils.pimsdateutil import hours_in_month, doytimestr_to_datetime, datestr_to_datetime
from pims.files.utils import mkdir_p, most_recent_file_with_suffix
from pims.database.samsquery import CuMonthlyQuery, _HOST, _SCHEMA, _UNAME, _PASSWD
from pims.excel.modification import overwrite_last_row_with_totals, kpi_sheet_fill
from openpyxl.reader.excel import load_workbook
from xlsxwriter.utility import xl_rowcol_to_cell, xl_range

############################################################################################
# PRELIMINARY STEPS:
#
# 1. Run EHS > Launchpad > NRT List Request for GMT day range, using:
#    -- hourly thinning = 3600
#    -- name convention for sto file = YYYY_MM_ddd_ddd_kpi.sto
# 
# 2. Copy sto file from TReK to /misc/yoda/www/plots/batch/padtimes/NRT_sto_files/
# 
# 3. Run following python script (NOTE: you need the one/only argument exactly as shown)
#    /home/pims/dev/programs/python/pims/pad/amp_kpi.py convert_latest_sto2xlsx
# 
# 4. The above creates sheet called "kpi" in following XLSX file:
#    /misc/yoda/www/plots/batch/padtimes/YYYY_MM_ddd_ddd_kpi.xlsx
#
############################################################################################
# THESE NEXT STEPS ARE CURRENTLY DONE MANUALLY (with hints for automation included):
#
# >> Use [openpyxl?] to fill in cells for "kpi" sheet in following steps...
# 
# 5. Get GMT range formatted string into kpi sheet cell B1
# 
# 6. Numerators gleaned from column headings
# 
# 7. Denominators gleaned from column headings
# 
# 8. Formatting


#  Group,                 System,     Resource,   Formula
#  ------------------------------------------------------------
#  continuous,            mams,       ossraw,     100 * D / T
#  continuous,            mams,       hirap,      100 * D / T
#  continuous,            sams,       121f03,     100 * D / T
#  continuous,            sams,       121f04,     100 * D / T
#  continuous,            sams,       cu,         100 * X / T
#  ------------------------------------------------------------
#  power_rack_dependent,  sams,       121f05,     100 * D / R
#  power_rack_dependent,  sams,       121f02,     100 * D / R
#  power_rack_dependent,  sams,       121f08,     100 * D / R
#  ------------------------------------------------------------
#  payload_dependent,     sams,       es03,       100 * D / P
#  payload_dependent,     sams,       es05,       100 * D / P
#  payload_dependent,     sams,       es06,       100 * D / P
#
# Each of below for a given resource for a given month (T is total hours for the given month)
# DONE continuous group, resource NOT "cu" = 100 * D / T; where D is the number of hours in PAD files
# TODO continuous group, resource IS  "cu" = 100 * X / T, where X is the number of records in the database in the CU Housekeeping Packet
# TODO power_rack_dependent group, involves MSID related to each ER power
# TODO payload_dependent group, resource IS "cir|fir", involves MSID related to SAMS power:
#      - UFZF12RT7452J is IOP MP PWR CTRL SAMS for FIR (ES06), at offset word 53 (byte 106), 1 byte length
#      - UFZF13RT7420J is IOP MP PWR CTRL SAMS for CIR (ES05), at offset word 53 (byte 106), 1 byte length
# TODO payload_dependent group, resource IS "msg", involves MSID related to outlet power (1 of 2 possible outlets):
#      - watching the MSG planning for which experiment is in, and then if they use SAMS, and then monitor the power outlet
#        to know when they turned us on

# filter and pivot to get aggregate sum of monthly hours
def monthly_hours(df, s):
    """filter and pivot to get aggregate sum of monthly hours"""
    ndf = df.filter(regex='Date|Year|Month|Day|' + s + '.*_hours')
    cols = [i for i in ndf.columns if i not in ['Date', 'Year', 'Month', 'Day']]
    t = pd.pivot_table(ndf, rows=['Year','Month'], values=cols, aggfunc=np.sum)
    series = t.transpose().sum()
    return series

# put systems' monthly hours (each a series) into pd.DataFrame
def monthly_hours_dataframe(df, systems_series):
    """put systems' monthly hours (each a series) into pd.DataFrame"""
    for k, v in systems_series.iteritems():
        systems_series[k] = monthly_hours(df, k)    
    monthly_hours_df = pd.DataFrame(systems_series)
    monthly_hours_df.columns = [ s.upper() for s in monthly_hours_df.columns ]
    return monthly_hours_df

# parse date/times from the csv
def parse(s):
    yr = int(s[0:4])
    #doy = int(s[5:8])
    mo = int(s[5:7])
    da = int(s[8:10])
    #hr = int(s[9:11])
    #min = int(s[12:14])
    #sec = int(s[15:17])
    #sec = float(sec)
    #mu_sec = int((sec - int(sec)) * 1e6)
    #mu_sec = 0
    #sec = int(sec)
    #dt = datetime(yr - 1, 12, 31)
    #delta = timedelta(days=doy, hours=hr, minutes=min, seconds=sec, microseconds=mu_sec)
    #return dt + delta
    return datetime.date(yr, mo, da)

# read big CSV into dataframe (for pivot tables)
def csv2dataframe(csvfile):
    """read CSV into dataframe (so we can use pivot tables)"""
    with open(csvfile, 'rb') as f:
        labels = f.next().strip().split(',')
        df = pd.read_csv( csvfile, parse_dates=True, index_col = [0] )
    return df

# read resource config CSV file into dataframe
def resource_csv2dataframe(csvfile):
    """read resource config CSV file into dataframe"""
    with open(csvfile, 'rb') as f:
        labels = f.next().strip().split(',')
        df = pd.read_csv( csvfile )
    return df

# generic normalize to get one/zero instead of on/off or closed/opened
def normalize_generic(v, one_list, zero_list):
    """generic normalize to get one/zero instead of on/off or closed/opened"""
    if isinstance(v, float) and np.isnan(v):
        return 0
    elif v.lower() in zero_list:
        return 0
    elif v.lower() in one_list:
        return 1
    else:
        return np.nan

# normalize on/off (as one/zero) for CIR and FIR power on/off columns
def normalize_cir_fir_power(v):
    """normalize on/off (as one/zero) for CIR and FIR power on/off columns"""
    if isinstance(v, float) and np.isnan(v):
        return 0
    elif v.lower().startswith('off') or v.lower().startswith('power off'):
        return 0
    elif v.lower().startswith('on') or v.lower().startswith('power on'):
        return 1
    else:
        return np.nan

# read NRT List Request output (sto, tab delimited) file into dataframe
def msg_cir_fir_sto2dataframe(stofile):
    """read ascii sto file into dataframe"""
    s = StringIO()
    with open(stofile, 'r') as f:
        # Read and ignore header lines
        header = f.readline() # labels
        s.write(header)
        is_data = False
        for line in f:
            if line.startswith('#Start_Data'):
                is_data = True
            if line.startswith('#End_Data'):
                is_data = False
            if is_data and not line.startswith('#Start_Data'):
                s.write(line)
    s.seek(0) # "rewind" to the beginning of the StringIO object
    df = pd.read_csv(s, sep='\t')
    
    # drop the unwanted "#Header" column
    df = df.drop('#Header', 1)
    column_labels = [ s.replace('Timestamp : Embedded GMT', 'GMT') for s in df.columns.tolist()]
    df.columns = column_labels
    
    # drop Unnamed columns
    for clabel in column_labels:
        if clabel.startswith('Unnamed'):
            df = df.drop(clabel, 1)

    # use Jen's nomenclature to rename column labels    
    msid_map = {
        'ULZL02RT0471C': 'MSG_Outlet_2_Current',
        'ULZL02RT0477J': 'MSG_Outlet_2_Status',
        'UFZF07RT0114V': 'MSG_Outlet1_28V',
        'UFZF07RT0118V': 'MSG_Outlet2_28V',
        'UFZF07RT0121J': 'MSG_Outlet1_Status',
        'UFZF07RT0125J': 'MSG_Outlet2_Status',
        'UFZF07RT0046T': 'MSW_WV_Air_Temp',
        'UFZF13RT7420J': 'TSH_ES05_CIR_Power_Status',
        'UFZF12RT7452J': 'TSH_ES06_FIR_Power_Status',
        'UEZE03RT1384C': 'ER3_Embedded_EE_Current',
        'UEZE03RT1548J': 'ER3_EE_F04_Power_Status',
        'UEZE04RT1394C': 'ER4_Drawer2_Current',
        'UEZE04RT1608J': 'ER4_Drawer2_Power_Status',
        'UEZE04RT1841J': 'ER4_Drawer2_Ethernet',
        }
    for k, v in msid_map.iteritems():
        df.rename(columns={k: v}, inplace=True)

    # normalize on/off (as one/zero) for CIR and FIR columns
    df.TSH_ES05_CIR_Power_Status = [ normalize_cir_fir_power(v) for v in df.TSH_ES05_CIR_Power_Status.values ]
    df.TSH_ES06_FIR_Power_Status = [ normalize_cir_fir_power(v) for v in df.TSH_ES06_FIR_Power_Status.values ]

    return df


# read NRT List Request output (sto, tab delimited) file into dataframe
def emcs_sto2dataframe(stofile):
    """read ascii sto file into dataframe"""
    
    s = StringIO()
    with open(stofile, 'r') as f:
        # Read and ignore header lines
        header = f.readline() # labels
        s.write(header)
        is_data = False
        for line in f:
            if line.startswith('#Start_Data'):
                is_data = True
            if line.startswith('#End_Data'):
                is_data = False
            if is_data and not line.startswith('#Start_Data'):
                s.write(line)
    s.seek(0) # "rewind" to the beginning of the StringIO object
    df = pd.read_csv(s, sep='\t')
    
    # drop the unwanted "#Header" column
    df = df.drop('#Header', 1)
    column_labels = [ s.replace('Timestamp : Embedded GMT', 'GMT') for s in df.columns.tolist()]
    df.columns = column_labels
    
    # drop Unnamed columns
    for clabel in column_labels:
        if clabel.startswith('Unnamed'):
            df = df.drop(clabel, 1)

    # use nomenclature to rename column labels    
    msid_map = {
            'UEZE08RT1001J': 'HS_Caution_Warning',
            'UEZE08RT1002J': 'HS_Cycle_Count',
            'UEZE08RT1003C': 'HS_ACS_Pump_Current',
            'UEZE08RT1004T': 'HS_ACS_TemperatureAcsSensorPEU',
            'UEZE08RT1005T': 'HS_ACS_TemperatureAcsSensorCEU',
            'UEZE08RT1006T': 'HS_TCS_TemperatureOTDAHX1',
            'UEZE08RT1007T': 'HS_TCS_TemperatureOTDAHX2',
            'UEZE08RT1008T': 'HS_TCS_TemperatureOTDAHX3',
            'UEZE08RT1009T': 'HS_TCS_TemperatureColdPlate1',
            'UEZE08RT1010T': 'HS_TCS_TemperatureColdPlate2',
            'UEZE08RT1011T': 'HS_TCS_TemperatureOTDPEU1',
            'UEZE08RT1012T': 'HS_TCS_TemperatureOTDPEU2',
            'UEZE08RT1013T': 'HS_TCS_TemperatureCEU',
            'UEZE08RT1014T': 'HS_TCS_TemperatureFan',
            'UEZE08RT1015T': 'HS_DA_Temperature1',
            'UEZE08RT1016T': 'HS_DA_Temperature2',
            'UEZE08RT1017T': 'HS_DB_Temperature1',
            'UEZE08RT1018T': 'HS_DB_Temperature2',
            'UEZE08RT1019T': 'HS_SPLC_Temperature',
            'UEZE08RT1020T': 'HS_VAS_Temperature',
            'UEZE08RT1021T': 'HS_VCR_Temperature',
            'UEZE08RT1022T': 'HS_DCDC_Temperature_EBox1',
            'UEZE08RT1023T': 'HS_DCDC_Temperature_EBox2',
            'UEZE08RT1024T': 'HS_DCDC_Temp RBLSS_Box1',
            'UEZE08RT1025T': 'HS_DCDC_Temp RBLSS_Box2',
            'UEZE08RT1026T': 'HS_EC1_Illumination_Temp',
            'UEZE08RT1027T': 'HS_EC2_Illumination_Temp',
            'UEZE08RT1028T': 'HS_EC3_Illumination_Temp',
            'UEZE08RT1029T': 'HS_EC4_Illumination_Temp',
            'UEZE08RT1030T': 'HS_DCDC_Temperature_Ebox1',
            'UEZE08RT1031T': 'HS_DCDC_Temperature_EBox2',
            'UEZE08RT1032T': 'HS_DCDC_Temp_RBLSS_Box1',
            'UEZE08RT1033T': 'HS_DCDC_Temp_RBLSS_Box2',
            'UEZE08RT1034T': 'HS_EC1_Illumination_Temp',
            'UEZE08RT1035T': 'HS_EC2_Illumination_Temp',
            'UEZE08RT1036T': 'HS_EC3_Illumination_Temp',
            'UEZE08RT1037T': 'HS_EC4_Illumination_Temp',
            'UEZE08RT1100J': 'Subset_ID',
            'UEZE08RT1101J': 'PEP_Service_Request_Word',
            'UEZE08RT1102J': 'PEP_Request_Supporting_Data',
            'UEZE03RT1005U': 'ER3_Data_Cycle_Counter',
            'UEZE03RT1389C': 'Unsign_28V_Output_18_Current_Locker3',
            'UEZE03RT1390C': 'Unsign_28V_Output_19_Current_Locker4',
            'UEZE03RT1391C': 'Unsign_28V_Output_20_Current_Locker7',
            'UEZE03RT1392C': 'Unsign_28V_Output_21_Current_Locker8',
            'UEZE03RT1393C': 'Unsign_28V_Output_22_Current_ISIS1)',
            'UEZE03RT1578J': 'Chan_18_OP_State',
            'UEZE03RT1584J': 'Chan_19_OP_State',
            'UEZE03RT1590J': 'Chan_20_OP_State',
            'UEZE03RT1596J': 'Chan_21_OP_State',
            'UEZE03RT1602J': 'Chan_22_OP_State',
            'UEZE03RT1840J': 'HRLC_PLD_Ether9_Active_Inactive',
            'LADP20MDJ764J': 'PL92_Active_Flag',
            'LADP20MDJ765J': 'PL92_HS_Data_Received'
        }
    for k, v in msid_map.iteritems():
        df.rename(columns={k: v}, inplace=True)
    
    ## normalize on/off (as one/zero) for CIR and FIR columns
    #df.TSH_ES05_CIR_Power_Status = [ normalize_cir_fir_power(v) for v in df.TSH_ES05_CIR_Power_Status.values ]
    #df.TSH_ES06_FIR_Power_Status = [ normalize_cir_fir_power(v) for v in df.TSH_ES06_FIR_Power_Status.values ]

    df.to_csv( stofile.replace('.sto', '.csv') )
    return df

#df = emcs_sto2dataframe('/misc/yoda/www/plots/user/handbook/source_docs/hb_vib_Columbus_EMCS_Candidates/EMCSdata.sto')
#raise SystemExit

# set diff as list
def list_diff(a, b):
    """set diff as list"""
    b = set(b)
    return [aa for aa in a if aa not in b]

# return subset of dataframe that have status == 'S'
def dataframe_subset(df, label, value_column, column_list):
    """return subset of dataframe that have status == 'S'"""
    # get and rename status column that corresponds to this value_column
    status_column = column_list[ column_list.index(value_column) + 1 ]

    # return empty dataframe subset if status column is all NaNs
    if np.all( pd.isnull( df[status_column] ) ):
        print 'status column %s is ALL NaNs for %s' % (status_column, label)
        df_sub = pd.DataFrame(columns=df.columns)
    # otherwise, we want rows with "S" as our subset
    else:
        df_sub = df[df[status_column] == 'S']

    # rename status column from like 'status.7' to like 'status.cir'
    new_status_column = 'status.' + label
    df_sub.rename(columns={status_column: new_status_column}, inplace=True)

    # drop the unwanted columns in brute force fashion
    for c in df_sub.columns:
        if c not in ['GMT', 'Date', value_column, new_status_column ]:
            df_sub = df_sub.drop(c, 1)

    # return dataframe subset for this label
    return df_sub

# process for specific payload (either CIR or FIR for now)
def process_cir_fir(df, label, value_column, column_list, stofile):
    """process for specific payload (either CIR or FIR for now)"""
    # new dataframe (subset) for this payload
    df_payload = dataframe_subset(df, label, value_column, column_list)

    # pivot to aggregate daily sum for "payload_hours" column
    grouped_payload = df_payload.groupby('Date').aggregate(np.sum)
    
    ## write payload info to CSVs
    #upcase = label.upper()
    #df_payload.to_csv( stofile.replace('.sto', '_' + upcase + '.csv') )
    #grouped_payload.to_csv( stofile.replace('.sto', '_' + upcase + '_grouped.csv') )
    
    # return dataframe for payload and date grouped dataframe too
    return df_payload, grouped_payload

# convert sto file to dataframe, then process and write to csv
def convert_sto2csv(stofile):
    """convert sto file to dataframe, then process and write to csv"""
    
    # get dataframe from sto file
    df = msg_cir_fir_sto2dataframe(stofile)
    column_list = df.columns.tolist()
    df.to_csv(stofile.replace('.sto', '_from_dataframe.csv'))
    
    # convert like 2014:077:00:02:04 to datetimes
    df['Date'] = [ doytimestr_to_datetime( doy_gmtstr ).date() for doy_gmtstr in df.GMT ]

    # convert datetimes to str and overwrite GMT with those strings
    df['GMT'] = [ d.strftime('%Y-%m-%d/%j,%H:%M:%S') for d in df.date ]

    # new dataframe (subset) for CIR
    df_cir, grouped_cir = process_cir_fir(df, 'cir', 'TSH_ES05_CIR_Power_Status', column_list, stofile)

    # new dataframe (subset) for FIR
    df_fir, grouped_fir = process_cir_fir(df, 'fir', 'TSH_ES06_FIR_Power_Status', column_list, stofile)
    
    # FIXME
    # - move zero_list and one_list up to here
    # - refactor commonanilty for ER3, ER4, MSG1, and MSG2
    
    # new dataframe (subset) for ER3 (ER3_EE_F04_Power_Status == 'CLOSED')
    df_er3 = dataframe_subset(df, 'er3', 'ER3_EE_F04_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    zero_list = ['off', 'power off', 'opened']
    one_list =  ['on' , 'power on' , 'closed']
    df_er3.ER3_EE_F04_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er3.ER3_EE_F04_Power_Status.values ]        
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_er3 = df_er3.groupby('Date').aggregate(np.sum)
    
    # write CSV for ER3
    df_er3.to_csv( stofile.replace('.sto', '_er3.csv') )
    grouped_er3.to_csv( stofile.replace('.sto', '_ER3_grouped.csv') )
    
    # new dataframe (subset) for ER4 (ER4_Drawer2_Power_Status == 'CLOSED')
    df_er4 = dataframe_subset(df, 'er4', 'ER4_Drawer2_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_er4.ER4_Drawer2_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er4.ER4_Drawer2_Power_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_er4 = df_er4.groupby('Date').aggregate(np.sum)    
    
    # write CSV for ER4
    df_er4.to_csv( stofile.replace('.sto', '_er4.csv') )
    grouped_er4.to_csv( stofile.replace('.sto', '_ER4_grouped.csv') )

    # new dataframe (subset) for MSG1 (MSG_Outlet1_Status == 'ON')
    df_msg1 = dataframe_subset(df, 'msg1', 'MSG_Outlet1_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg1.MSG_Outlet1_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg1.MSG_Outlet1_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg1 = df_msg1.groupby('Date').aggregate(np.sum)    
    
    # write CSV for MSG1
    df_msg1.to_csv( stofile.replace('.sto', '_msg1.csv') )
    grouped_msg1.to_csv( stofile.replace('.sto', '_MSG1_grouped.csv') )

    # new dataframe (subset) for MSG2 (MSG_Outlet2_Status == 'ON')
    df_msg2 = dataframe_subset(df, 'msg2', 'MSG_Outlet2_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg2.MSG_Outlet2_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg2.MSG_Outlet2_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg2 = df_msg2.groupby('Date').aggregate(np.sum)    
    
    # write CSV for MSG2
    df_msg2.to_csv( stofile.replace('.sto', '_msg2.csv') )
    grouped_msg2.to_csv( stofile.replace('.sto', '_MSG2_grouped.csv') )

# get dataframe with padtimes aggregate
def get_padtimes_aggregate():
    """get dataframe with padtimes aggregate"""
    pass

def demo_modify_existing():
    wb = load_workbook(filename = r'/tmp/empty_book.xlsx')
    ws = wb.get_sheet_by_name("ER3raw")
    #ws.cell(row=0, column=0).delete()
    print ws.cell('A19').value
    #ws.cell('A19').value = '/</-/'
    #wb.save(filename = r'/tmp/empty_book.xlsx')

# convert most recent sto file to dataframe, then process and write to xlsx
def convert_latest_sto2xlsx():
    """convert most recent sto file to dataframe, then process and write to xlsx"""
    
    # get most recent sto file along default dir
    stofile = most_recent_file_with_suffix('/misc/yoda/www/plots/batch/padtimes/NRT_sto_files', '.sto')

    # check if xlsx exists already
    xlsxfile = stofile.replace('.sto', '.xlsx').replace('padtimes/NRT_sto_files','padtimes')
    if os.path.exists( xlsxfile ):
        print "Abort: already have XLSX file %s for\nthe latest stofile = %s" % (xlsxfile, stofile)
        return

    # at this point, xlsx version does not exist, so convert most recent sto to xlsx
    convert_sto2xlsx(stofile, xlsxfile)

# read config info from xlsx file into dataframe
def read_config_template(xlsx_file='/misc/yoda/www/plots/batch/padtimes/amp_kpi_config.xlsx'):
    xl_file = pd.ExcelFile(xlsx_file)
    #dfs = {sheet_name: xl_file.parse(sheet_name) for sheet_name in xl_file.sheet_names}
    df_config = xl_file.parse('config')
    return df_config

# convert sto file to dataframe, then process and write to xlsx
def convert_sto2xlsx(stofile, xlsxfile):
    """convert sto file to dataframe, then process and write to xlsx"""
    
    # get dataframe from sto file
    df = msg_cir_fir_sto2dataframe(stofile)
    column_list = df.columns.tolist()
    #df.to_csv(stofile.replace('.sto', '_from_dataframe.csv'))
    
    # convert like 2014:077:00:02:04 to datetimes
    df['Date'] = [ doytimestr_to_datetime( doy_gmtstr ).date() for doy_gmtstr in df.GMT ]

    # get date range from sto file
    date_min = min(df['Date'])
    date_max = max(df['Date'])

    # convert datetimes to str and overwrite GMT with those strings
    df['GMT'] = [ d.strftime('%Y-%m-%d/%j,%H:%M:%S') for d in df.Date ]

    ### CIR ###
    # new dataframe (subset) for CIR
    df_cir, grouped_cir = process_cir_fir(df, 'cir', 'TSH_ES05_CIR_Power_Status', column_list, stofile)

    ### FIR ###
    # new dataframe (subset) for FIR
    df_fir, grouped_fir = process_cir_fir(df, 'fir', 'TSH_ES06_FIR_Power_Status', column_list, stofile)
    
    # FIXME
    # - move zero_list and one_list up to here
    # - refactor commonanilty for ER3, ER4, MSG1, and MSG2
    
    ### ER3 ###
    # new dataframe (subset) for ER3 (ER3_EE_F04_Power_Status == 'CLOSED')
    df_er3 = dataframe_subset(df, 'er3', 'ER3_EE_F04_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    zero_list = ['off', 'power off', 'opened']
    one_list =  ['on' , 'power on' , 'closed']
    df_er3.ER3_EE_F04_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er3.ER3_EE_F04_Power_Status.values ]        
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_er3 = df_er3.groupby('Date').aggregate(np.sum)
    
    ### ER4 ###    
    # new dataframe (subset) for ER4 (ER4_Drawer2_Power_Status == 'CLOSED')
    df_er4 = dataframe_subset(df, 'er4', 'ER4_Drawer2_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_er4.ER4_Drawer2_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er4.ER4_Drawer2_Power_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_er4 = df_er4.groupby('Date').aggregate(np.sum)    

    ### MSG1 ###
    # new dataframe (subset) for MSG1 (MSG_Outlet1_Status == 'ON')
    df_msg1 = dataframe_subset(df, 'msg1', 'MSG_Outlet1_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg1.MSG_Outlet1_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg1.MSG_Outlet1_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg1 = df_msg1.groupby('Date').aggregate(np.sum)    

    ### MSG2 ###
    # new dataframe (subset) for MSG2 (MSG_Outlet2_Status == 'ON')
    df_msg2 = dataframe_subset(df, 'msg2', 'MSG_Outlet2_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg2.MSG_Outlet2_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg2.MSG_Outlet2_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg2 = df_msg2.groupby('Date').aggregate(np.sum)    

    # get dataframe with CU hour count info
    cu = CuMonthlyQuery(_HOST, _SCHEMA, _UNAME, _PASSWD, date_min, date_max)
    cu_querystr = cu._get_query(date_min, date_max)
    s = StringIO()
    s.write(str(cu))
    s.seek(0) # "rewind" to the beginning of the StringIO object
    df_cu = pd.read_csv(s, sep='\t', parse_dates=True, index_col = [0])
    
    # get dataframe for padtimes
    df_tmp = pd.read_csv('/misc/yoda/www/plots/batch/padtimes/padtimes.csv', parse_dates=True, index_col = [0])
    df_pad = df_tmp.filter(regex='Date|.*_hours')
    
    # create wall clock dataframe
    df_wall_clock = grouped_er3.copy(deep=True)

    # rename column to Wall_Clock
    df_wall_clock.rename(columns={'ER3_EE_F04_Power_Status':  'Wall_Clock'}, inplace=True)
    df_wall_clock['Wall_Clock'] = 24
    
    # merge dataframes and just add those on to df_pad
    bamf_df = df_wall_clock.join([grouped_er3, grouped_er4, grouped_cir, grouped_fir, grouped_msg1, grouped_msg2, df_cu, df_pad])

    # drop unwanted columns
    unwanted_columns = ['mams_ossbtmf_hours', 'iss_radgse_hours', 'mma_0bba_hours', 'mma_0bbb_hours', 'mma_0bbc_hours', 'mma_0bbd_hours']
    for uc in unwanted_columns:
        bamf_df = bamf_df.drop(uc, 1)    

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(xlsxfile, engine='xlsxwriter')

    # Create kpi sheet for monthly percentages
    df_config = read_config_template()
    df_config.to_excel(writer, sheet_name='kpi', index=False)
    
    # Create sheets for dataframes
    bamf_df.to_excel(writer, sheet_name='raw', index=True)
    
    # Close the Pandas Excel writer and output the Excel file.
    writer.save()
    print 'wrote raw and kpi template sheets to %s' % xlsxfile

    # Reckon GMT range and overwrite last row with totals; where last row is day 1 of next month
    # and write GMT range into kpi sheet cell B1
    overwrite_last_row_with_totals(xlsxfile, df_config, bamf_df)
    print 'reckoned GMT range\nmodified and saved %s with formulas and totals' % xlsxfile
    
# produce output csv with per-system monthly sensor hour totals
def main(csvfile, resource_csvfile):
    """produce output csv with per-system monthly sensor hour totals"""
    
    # read resource config csv file
    df_cfg = resource_csv2dataframe(resource_csvfile)
    regex_sensor_hours = '.*' + '_hours|.*'.join(df_cfg['Resource']) + '_hours'
    
    #print df_cfg; raise SystemExit
    
    # read input CSV into big pd.DataFrame
    df = csv2dataframe(csvfile)

    # filter to keep only hours columns (gets rid of bytes columns) for each sensor
    # that shows up in df_cfg's Resource column
    ndf = df.filter(regex='Date|Year|Month|Day|' + regex_sensor_hours)
    
    # pivot to aggregate monthly sum for each "sensor_hours" column
    t = pd.pivot_table(ndf, rows=['Year','Month'], aggfunc=np.sum)
    
    # drop the unwanted "Day" column
    df_monthly_hours = t.drop('Day', 1)
    
    # convert index, which are tuples like (YEAR, MONTH), to tuples like (YEAR, MONTH, 1)
    date_tuples = [ ( t + (1,) ) for t in df_monthly_hours.index.values ]

    # convert date_tuples each to hours_in_month
    hours = [ hours_in_month( datetime.date( *tup ) ) for tup in date_tuples ]

    # before we add hours_in_month column, get list of columns for iteration below
    cols = df_monthly_hours.columns.tolist()
    
    # now we can append month_hours column
    df_monthly_hours['hours_in_month'] = pd.Series( hours, index=df_monthly_hours.index)
    
    # iterate over columns (above code helps exclude hours_in_month) to get 100*sensor_hours/hours_in_month
    for c in cols:
        pctstr = c + '_pct'
        pct = 100 * df_monthly_hours[c] / df_monthly_hours['hours_in_month']
        df_monthly_hours[pctstr] = pd.Series( np.round(pct, decimals=0), index=df_monthly_hours.index)
    
    # save csv output file
    csvout = csvfile.replace('.csv','_monthly_hours.csv')
    df_monthly_hours.to_csv(csvout)
    print 'wrote %s' % csvout

if __name__ == '__main__':
    
    if len(sys.argv) == 2:
        # special run, manually like: amp_kpi.py convert_latest_sto2xlsx
        eval( sys.argv[1] + '()' )
        raise SystemExit
        
    elif len(sys.argv) == 3:
        # special input arg file locations (see else clause below)
        csvfile = sys.argv[1]
        resource_csvfile = sys.argv[2]
        
    else:
        # typical run via cronjob
        csvfile = '/misc/yoda/www/plots/batch/padtimes/padtimes.csv'
        resource_csvfile = '/misc/yoda/www/plots/batch/padtimes/kpi_track.csv'
    
    main(csvfile, resource_csvfile)
