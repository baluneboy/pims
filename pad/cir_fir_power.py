#!/usr/bin/env python

import re
import os
import sys
import csv
import numpy as np
import scipy.io as spio
import pandas as pd
import datetime
from cStringIO import StringIO

from pims.utils.pimsdateutil import hours_in_month, doytimestr_to_datetime, datestr_to_datetime
from pims.files.utils import mkdir_p, most_recent_file_with_suffix
from pims.database.samsquery import CuMonthlyQuery, _HOST_SAMS, _SCHEMA_SAMS, _UNAME_SAMS, _PASSWD_SAMS
from pims.excel.modification import overwrite_last_row_with_totals, kpi_sheet_fill
from openpyxl.reader.excel import load_workbook
from xlsxwriter.utility import xl_rowcol_to_cell, xl_range

############################################################################################
# PRELIMINARY STEPS:
#
# 1. Run EHS > Launchpad > NRT List Request for GMT day range, using:
#    -- hourly thinning = 3600
#    -- name convention for sto file = YYYY_MM_ddd_ddd_kpi.sto
# 
# 2. Copy sto file from TReK to /misc/yoda/www/plots/batch/padtimes/NRT_sto_files/
# 
# 3. Run following python script (NOTE: you need the one/only argument exactly as shown)
#    /home/pims/dev/programs/python/pims/pad/amp_kpi.py convert_latest_sto2xlsx
# 
# 4. The above creates sheet called "kpi" in following XLSX file:
#    /misc/yoda/www/plots/batch/padtimes/YYYY_MM_ddd_ddd_kpi.xlsx
#
############################################################################################
# THESE NEXT STEPS ARE CURRENTLY DONE MANUALLY (with hints for automation included):
#
# >> Use [openpyxl?] to fill in cells for "kpi" sheet in following steps...
# 
# 5. Get GMT range formatted string into kpi sheet cell B1
# 
# 6. Numerators gleaned from column headings
# 
# 7. Denominators gleaned from column headings
# 
# 8. Formatting


MY_MSID_MAP = {
        'LAPR33FC1017J': 'CIR_LABS3_Pwr_Hours',
        'LAPR31FC1017J': 'FIR_LABS4_Pwr_Hours',
        # 'UFZF12RT7452J': 'TSH_ES06_FIR_Power_Status',
        # 'UFZF12RT7452J': 'TSH_ES06_FIR_Power_Status',
        # 'UFZF13RT7420J':  'IOP_MP_PWR_CTRL_SAMS'
}


def monthly_hours(df, s):
    """filter and pivot to get aggregate sum of monthly hours"""
    ndf = df.filter(regex='Date|Year|Month|Day|' + s + '.*_hours')
    cols = [i for i in ndf.columns if i not in ['Date', 'Year', 'Month', 'Day']]
    t = pd.pivot_table(ndf, rows=['Year','Month'], values=cols, aggfunc=np.sum)
    series = t.transpose().sum()
    return series


def monthly_hours_dataframe(df, systems_series):
    """put systems' monthly hours (each a series) into pd.DataFrame"""
    for k, v in systems_series.iteritems():
        systems_series[k] = monthly_hours(df, k)    
    monthly_hours_df = pd.DataFrame(systems_series)
    monthly_hours_df.columns = [ s.upper() for s in monthly_hours_df.columns ]
    return monthly_hours_df


def parse(s):
    yr = int(s[0:4])
    #doy = int(s[5:8])
    mo = int(s[5:7])
    da = int(s[8:10])
    #hr = int(s[9:11])
    #min = int(s[12:14])
    #sec = int(s[15:17])
    #sec = float(sec)
    #mu_sec = int((sec - int(sec)) * 1e6)
    #mu_sec = 0
    #sec = int(sec)
    #dt = datetime(yr - 1, 12, 31)
    #delta = timedelta(days=doy, hours=hr, minutes=min, seconds=sec, microseconds=mu_sec)
    #return dt + delta
    return datetime.date(yr, mo, da)


def normalize_cir_fir_power(v):
    """normalize on/off (as one/zero) for CIR and FIR power on/off columns"""
    if isinstance(v, float) and np.isnan(v):
        return 0
    elif v.lower().startswith('off') or v.lower().startswith('power off') or v.lower().startswith('open'):
        return 0
    elif v.lower().startswith('on') or v.lower().startswith('power on') or v.lower().startswith('closed'):
        return 1
    else:
        return np.nan


def show_len(df):
    print 'dataframe has length = %d' % len(df)


def generic_sto2dataframe(stofile, query_chain=None, predicate=None):
    """read ascii sto file into dataframe"""
    s = StringIO()
    with open(stofile, 'r') as f:
        # Read and ignore header lines
        header = f.readline() # labels
        s.write(header)
        is_data = False
        for line in f:
            if line.startswith('#Start_Data'):
                is_data = True
            if line.startswith('#End_Data'):
                is_data = False
            if is_data and not line.startswith('#Start_Data'):
                s.write(line)
    s.seek(0) # "rewind" to the beginning of the StringIO object
    df = pd.read_csv(s, sep='\t')
    
    # drop the unwanted "#Header" column
    df = df.drop('#Header', 1)
    column_labels = [ s.replace('Timestamp : Embedded GMT', 'GMT') for s in df.columns.tolist()]
    df.columns = column_labels
    
    # drop Unnamed columns
    for clabel in column_labels:
        if clabel.startswith('Unnamed'):
            df = df.drop(clabel, 1)

    # FIXME verify these mappings, especially for MSG...
    # use our nomenclature to rename column labels    
    msid_map = {
        'ULZL02RT0471C': 'MSG_Outlet_2_Current',
        'ULZL02RT0477J': 'MSG_Outlet_2_Status',
        'UFZF07RT0114V': 'MSG_Outlet1_28V',
        'UFZF07RT0118V': 'MSG_Outlet2_28V',
        'UFZF07RT0121J': 'MSG_Outlet1_Status',
        'UFZF07RT0125J': 'MSG_Outlet2_Status',
        'UFZF07RT0046T': 'MSW_WV_Air_Temp',
        'UFZF13RT7420J': 'TSH_ES05_CIR_Power_Status',
        'UFZF12RT7452J': 'TSH_ES06_FIR_Power_Status',
        'UEZE03RT1384C': 'ER3_Embedded_EE_Current',
        'UEZE03RT1548J': 'ER3_EE_F04_Power_Status',
        'UEZE05RT1394C': 'ER5_Drawer2_Current',
        'UEZE05RT1608J': 'ER5_Drawer2_Power_Status',
        'UEZE05RT1841J': 'ER5_Drawer2_Ethernet',
        'UEZE06RT1578J': 'ER6_Locker3_Status',
        'UEZE06RT1389C': 'ER6_Locker3_Current',
        'UEZE06RT1584J': 'TSH_ES20_HER_Power_Status',
        'UEZE06RT1390C': 'ER6_Locker4_Current',
        'UEZE12RT1384C': 'ER7_Embedded_EE_Current',  # added on 2018-08-23
        'UEZE12RT1548J': 'ER7_EE_F01_Power_Status',  # added on 2018-08-23        
        }
    for k, v in msid_map.iteritems():
        df.rename(columns={k: v}, inplace=True)

    # apply query chain if it's not None
    if query_chain:
        for query_str in query_chain:
            df = df.query(query_str)

    # apply predicate to resultant dataframe if it's not None
    if predicate:
        predicate(df)
        
    return df


def msg_cir_fir_sto2dataframe(stofile):
    """read ascii sto file into dataframe for MSG/CIR/FIR"""
    s = StringIO()
    with open(stofile, 'r') as f:
        # Read and ignore header lines
        header = f.readline() # labels
        s.write(header)
        is_data = False
        for line in f:
            if line.startswith('#Start_Data'):
                is_data = True
            if line.startswith('#End_Data'):
                is_data = False
            if is_data and not line.startswith('#Start_Data'):
                s.write(line)
    s.seek(0)  # "rewind" to the beginning of the StringIO object
    df = pd.read_csv(s, sep='\t')
    
    # drop the unwanted "#Header" column
    df = df.drop('#Header', 1)
    column_labels = [ s.replace('Timestamp : Embedded GMT', 'GMT') for s in df.columns.tolist()]
    df.columns = column_labels
    
    # drop Unnamed columns
    for clabel in column_labels:
        if clabel.startswith('Unnamed'):
            df = df.drop(clabel, 1)

    """
         MSID             Technical Name
        --------------------------------------------
        'ULZL02RT0471C': 'ESEM 4B OUTLET 2 CURRENT',
        'ULZL02RT0477J': 'ESEM 4B OUTLET 2 STATUS',
        'UFZF07RT0114V': '+28V OUTLET 1',
        'UFZF07RT0118V': '+28V OUTLET 2',
        'UFZF07RT0121J': '+28V OUTLET 1 ON',
        'UFZF07RT0125J': '+28V OUTLET 2 ON',
        'UFZF07RT0046T': 'TEMP INLET',
        'UFZF13RT7420J': 'IOP_MP_PWR_CTRL_SAMS',
        'UFZF12RT7452J': 'IOP_MP_PWR_CTRL_SAMS',
        'UEZE03RT1384C': 'SSPCM O/P 14',
        'UEZE03RT1548J': 'SSPCM CONF (T4-63) 68',
        'UEZE05RT1394C': 'SSPCM O/P 24'.
        'UEZE05RT1608J': 'SSPCM CONF (T4-63) 118',
        'UEZE05RT1841J': 'HRLC BIT 21',
        'UEZE06RT1578J': 'SSPCM CONF (T4-63) 93',
        'UEZE06RT1389C': 'SSPCM O/P 19',        
    """

    # FIXME verify these mappings, especially for MSG...
    # use our nomenclature to rename column labels    
    msid_map = {
        'ULZL02RT0471C': 'MSG_Outlet_2_Current',
        'ULZL02RT0477J': 'MSG_Outlet_2_Status',
        'UFZF07RT0114V': 'MSG_Outlet1_28V',
        'UFZF07RT0118V': 'MSG_Outlet2_28V',
        'UFZF07RT0121J': 'MSG_Outlet1_Status',
        'UFZF07RT0125J': 'MSG_Outlet2_Status',
        'UFZF07RT0046T': 'MSW_WV_Air_Temp',
        'UFZF13RT7420J': 'TSH_ES05_CIR_Power_Status',
        'UFZF12RT7452J': 'TSH_ES06_FIR_Power_Status',
        'UEZE03RT1384C': 'ER3_Embedded_EE_Current',
        'UEZE03RT1548J': 'ER3_EE_F04_Power_Status',
        'UEZE05RT1394C': 'ER5_Drawer2_Current',
        'UEZE05RT1608J': 'ER5_Drawer2_Power_Status',
        'UEZE05RT1841J': 'ER5_Drawer2_Ethernet',
        'UEZE06RT1578J': 'ER6_Locker3_Status',
        'UEZE06RT1389C': 'ER6_Locker3_Current',
        'UEZE06RT1584J': 'TSH_ES20_HER_Power_Status',
        'UEZE06RT1390C': 'ER6_Locker4_Current',
        'UEZE12RT1384C': 'ER7_Embedded_EE_Current',  # added on 2018-08-23
        'UEZE12RT1548J': 'ER7_EE_F01_Power_Status',  # added on 2018-08-23        
        }
    for k, v in msid_map.iteritems():
        df.rename(columns={k: v}, inplace=True)

    # normalize on/off (as one/zero) for CIR and FIR columns
    df.TSH_ES05_CIR_Power_Status = [ normalize_cir_fir_power(v) for v in df.TSH_ES05_CIR_Power_Status.values ]
    df.TSH_ES06_FIR_Power_Status = [ normalize_cir_fir_power(v) for v in df.TSH_ES06_FIR_Power_Status.values ]
    df.TSH_ES20_HER_Power_Status = [ normalize_cir_fir_power(v) for v in df.TSH_ES20_HER_Power_Status.values ]

    return df


def list_diff(a, b):
    """set diff as list"""
    b = set(b)
    return [aa for aa in a if aa not in b]


def dataframe_subset(df, label, value_column, column_list):
    """return subset of dataframe that have status == 'S'"""
    # get and rename status column that corresponds to this value_column
    status_column = column_list[column_list.index(value_column) + 1]

    # return empty dataframe subset if status column is all NaNs
    if np.all( pd.isnull( df[status_column] ) ):
        print 'status column %s is ALL NaNs for %s' % (status_column, label)
        df_sub = pd.DataFrame(columns=df.columns)
    # otherwise, we want rows with "S" as our subset
    else:
        df_sub = df[df[status_column] == 'S']

    # rename status column from like 'status.7' to like 'status.cir'
    new_status_column = 'status.' + label
    #df_sub.rename(columns={status_column: new_status_column}, inplace=True)
    df_sub = df_sub.rename(columns={status_column: new_status_column})

    # drop the unwanted columns in brute force fashion
    for c in df_sub.columns:
        if c not in ['GMT', 'Date', value_column, new_status_column ]:
            df_sub = df_sub.drop(c, 1)

    # return dataframe subset for this label
    return df_sub


def process_cir_fir(df, label, value_column, column_list, stofile):
    """process for specific payload (either CIR or FIR for now)"""
    # new dataframe (subset) for this payload
    df_payload = dataframe_subset(df, label, value_column, column_list)

    # pivot to aggregate daily sum for "payload_hours" column
    # grouped_payload = df_payload.groupby('Date').aggregate(np.sum)
    grouped_payload = df_payload.groupby('GMT').aggregate(np.sum)

    ## write payload info to CSVs
    #upcase = label.upper()
    #df_payload.to_csv( stofile.replace('.sto', '_' + upcase + '.csv') )
    #grouped_payload.to_csv( stofile.replace('.sto', '_' + upcase + '_grouped.csv') )
    
    # return dataframe for payload and date grouped dataframe too
    return df_payload, grouped_payload


def convert_sto2csv(stofile):
    """convert sto file to dataframe, then process and write to csv"""
    
    # get dataframe from sto file
    df = msg_cir_fir_sto2dataframe(stofile)
    column_list = df.columns.tolist()
    df.to_csv(stofile.replace('.sto', '_from_dataframe.csv'))
    
    # convert like 2014:077:00:02:04 to datetimes
    df['Date'] = [ doytimestr_to_datetime( doy_gmtstr ).date() for doy_gmtstr in df.GMT ]

    # convert datetimes to str and overwrite GMT with those strings
    df['GMT'] = [ d.strftime('%Y-%m-%d/%j,%H:%M:%S') for d in df.date ]

    # new dataframe (subset) for CIR
    df_cir, grouped_cir = process_cir_fir(df, 'cir', 'TSH_ES05_CIR_Power_Status', column_list, stofile)

    # new dataframe (subset) for FIR
    df_fir, grouped_fir = process_cir_fir(df, 'fir', 'TSH_ES06_FIR_Power_Status', column_list, stofile)

    # new dataframe (subset) for Hermes
    df_hermes, grouped_hermes = process_hermes(df, 'hermes', 'TSH_ES20_HER_Power_Status', column_list, stofile)

    # FIXME
    # - move zero_list and one_list up to here
    # - refactor commonanilty for ER3, ER4, MSG1, and MSG2
    
    # new dataframe (subset) for ER3 (ER3_EE_F04_Power_Status == 'CLOSED')
    df_er3 = dataframe_subset(df, 'er3', 'ER3_EE_F04_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    zero_list = ['off', 'power off', 'opened']
    one_list =  ['on' , 'power on' , 'closed']
    df_er3.ER3_EE_F04_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er3.ER3_EE_F04_Power_Status.values ]        
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_er3 = df_er3.groupby('Date').aggregate(np.sum)
    
    # write CSV for ER3
    df_er3.to_csv( stofile.replace('.sto', '_er3.csv') )
    grouped_er3.to_csv( stofile.replace('.sto', '_ER3_grouped.csv') )
    
    # new dataframe (subset) for ER5 (ER5_Drawer2_Power_Status == 'CLOSED')
    df_er5 = dataframe_subset(df, 'er5', 'ER5_Drawer2_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_er5.ER5_Drawer2_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er5.ER5_Drawer2_Power_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_er5 = df_er5.groupby('Date').aggregate(np.sum)    
    
    # write CSV for ER5
    df_er5.to_csv( stofile.replace('.sto', '_er5.csv') )
    grouped_er5.to_csv( stofile.replace('.sto', '_ER5_grouped.csv') )

    # new dataframe (subset) for MSG1 (MSG_Outlet1_Status == 'ON')
    df_msg1 = dataframe_subset(df, 'msg1', 'MSG_Outlet1_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg1.MSG_Outlet1_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg1.MSG_Outlet1_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg1 = df_msg1.groupby('Date').aggregate(np.sum)    
    
    # write CSV for MSG1
    df_msg1.to_csv( stofile.replace('.sto', '_msg1.csv') )
    grouped_msg1.to_csv( stofile.replace('.sto', '_MSG1_grouped.csv') )

    # new dataframe (subset) for MSG2 (MSG_Outlet2_Status == 'ON')
    df_msg2 = dataframe_subset(df, 'msg2', 'MSG_Outlet2_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg2.MSG_Outlet2_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg2.MSG_Outlet2_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg2 = df_msg2.groupby('Date').aggregate(np.sum)    
    
    # write CSV for MSG2
    df_msg2.to_csv( stofile.replace('.sto', '_msg2.csv') )
    grouped_msg2.to_csv( stofile.replace('.sto', '_MSG2_grouped.csv') )
    """convert sto file to dataframe, then process and write to csv"""
    
    # get dataframe from sto file
    df = msg_cir_fir_sto2dataframe(stofile)
    column_list = df.columns.tolist()
    df.to_csv(stofile.replace('.sto', '_from_dataframe.csv'))
    
    # convert like 2014:077:00:02:04 to datetimes
    df['Date'] = [ doytimestr_to_datetime( doy_gmtstr ).date() for doy_gmtstr in df.GMT ]

    # convert datetimes to str and overwrite GMT with those strings
    df['GMT'] = [ d.strftime('%Y-%m-%d/%j,%H:%M:%S') for d in df.date ]

    # new dataframe (subset) for CIR
    df_cir, grouped_cir = process_cir_fir(df, 'cir', 'TSH_ES05_CIR_Power_Status', column_list, stofile)

    # new dataframe (subset) for FIR
    df_fir, grouped_fir = process_cir_fir(df, 'fir', 'TSH_ES06_FIR_Power_Status', column_list, stofile)

    # new dataframe (subset) for Hermes
    df_hermes, grouped_hermes = process_hermes(df, 'hermes', 'TSH_ES20_HER_Power_Status', column_list, stofile)

    # FIXME
    # - move zero_list and one_list up to here
    # - refactor commonanilty for ER3, ER4, MSG1, and MSG2
    
    # new dataframe (subset) for ER3 (ER3_EE_F04_Power_Status == 'CLOSED')
    df_er3 = dataframe_subset(df, 'er3', 'ER3_EE_F04_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    zero_list = ['off', 'power off', 'opened']
    one_list =  ['on' , 'power on' , 'closed']
    df_er3.ER3_EE_F04_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er3.ER3_EE_F04_Power_Status.values ]        
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_er3 = df_er3.groupby('Date').aggregate(np.sum)
    
    # write CSV for ER3
    df_er3.to_csv( stofile.replace('.sto', '_er3.csv') )
    grouped_er3.to_csv( stofile.replace('.sto', '_ER3_grouped.csv') )
    
    # new dataframe (subset) for ER5 (ER5_Drawer2_Power_Status == 'CLOSED')
    df_er5 = dataframe_subset(df, 'er5', 'ER5_Drawer2_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_er5.ER5_Drawer2_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er5.ER5_Drawer2_Power_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_er5 = df_er5.groupby('Date').aggregate(np.sum)    
    
    # write CSV for ER5
    df_er5.to_csv( stofile.replace('.sto', '_er5.csv') )
    grouped_er5.to_csv( stofile.replace('.sto', '_ER5_grouped.csv') )

    # new dataframe (subset) for MSG1 (MSG_Outlet1_Status == 'ON')
    df_msg1 = dataframe_subset(df, 'msg1', 'MSG_Outlet1_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg1.MSG_Outlet1_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg1.MSG_Outlet1_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg1 = df_msg1.groupby('Date').aggregate(np.sum)    
    
    # write CSV for MSG1
    df_msg1.to_csv( stofile.replace('.sto', '_msg1.csv') )
    grouped_msg1.to_csv( stofile.replace('.sto', '_MSG1_grouped.csv') )

    # new dataframe (subset) for MSG2 (MSG_Outlet2_Status == 'ON')
    df_msg2 = dataframe_subset(df, 'msg2', 'MSG_Outlet2_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg2.MSG_Outlet2_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg2.MSG_Outlet2_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg2 = df_msg2.groupby('Date').aggregate(np.sum)    
    
    # write CSV for MSG2
    df_msg2.to_csv( stofile.replace('.sto', '_msg2.csv') )
    grouped_msg2.to_csv( stofile.replace('.sto', '_MSG2_grouped.csv') )


def get_padtimes_aggregate():
    """get dataframe with padtimes aggregate"""
    pass


def demo_modify_existing():
    wb = load_workbook(filename = r'/tmp/empty_book.xlsx')
    ws = wb.get_sheet_by_name("ER3raw")
    #ws.cell(row=0, column=0).delete()
    print ws.cell('A19').value
    #ws.cell('A19').value = '/</-/'
    #wb.save(filename = r'/tmp/empty_book.xlsx')


def convert_latest_sto2xlsx():
    """convert most recent sto file to dataframe, then process and write to xlsx"""
    
    # get most recent sto file along default dir
    stofile = most_recent_file_with_suffix('/misc/yoda/www/plots/batch/padtimes/NRT_sto_files', '.sto')

    print 'STO file is: ', stofile
    
    # check if xlsx exists already
    xlsxfile = stofile.replace('.sto', '.xlsx').replace('padtimes/NRT_sto_files','padtimes')
    if os.path.exists( xlsxfile ):
        print "Abort: already have XLSX file %s for\nthe latest stofile = %s" % (xlsxfile, stofile)
        return

    # at this point, xlsx version does not exist, so convert most recent sto to xlsx
    convert_sto2xlsx(stofile, xlsxfile)


def read_config_template(xlsx_file='/misc/yoda/www/plots/batch/padtimes/amp_kpi_config.xlsx'):
    """read config info from xlsx file into dataframe"""
    xl_file = pd.ExcelFile(xlsx_file)
    #dfs = {sheet_name: xl_file.parse(sheet_name) for sheet_name in xl_file.sheet_names}
    df_config = xl_file.parse('config')
    return df_config


def convert_sto2csv(stofile, msid_map=MY_MSID_MAP):
    """convert sto file to dataframe"""
    
    # get dataframe from sto file
    df = sto2dataframe(stofile, msid_map)
    column_list = df.columns.tolist()
    df.to_csv(stofile.replace('.sto', '.sto.csv'))


def convert_sto2xlsx(stofile, xlsxfile):
    """convert sto file to dataframe, then process and write to xlsx"""
    
    # get dataframe from sto file
    df = msg_cir_fir_sto2dataframe(stofile)
    column_list = df.columns.tolist()
    #df.to_csv(stofile.replace('.sto', '_from_dataframe.csv'))

    # print(df)
    # print(column_list)
    # print('hey')
    
    # convert like 2014:077:00:02:04 to dates
    df['Date'] = [ doytimestr_to_datetime( doy_gmtstr ).date() for doy_gmtstr in df.GMT ]
    df['dtm'] = [ doytimestr_to_datetime( doy_gmtstr ) for doy_gmtstr in df.GMT ]

    # get date range from sto file
    date_min = min(df['Date'])
    date_max = max(df['Date'])

    # convert datetimes to str and overwrite GMT with those strings
    df['GMT'] = [ d.strftime('%Y-%m-%d/%j,%H:%M:%S') for d in df.dtm ]

    ################################
    ### CIR ###
    # new dataframe (subset) for CIR with grouped_cir being sum of payload hours on daily basis
    df_cir, grouped_cir = process_cir_fir(df, 'cir', 'TSH_ES05_CIR_Power_Status', column_list, stofile)

    ################################
    ### FIR ###
    # new dataframe (subset) for FIR with grouped_fir being sum of payload hours on daily basis
    df_fir, grouped_fir = process_cir_fir(df, 'fir', 'TSH_ES06_FIR_Power_Status', column_list, stofile)

    ################################
    ### Hermes ###
    # new dataframe (subset) for Hermes with grouped_fir being sum of payload hours on daily basis
    df_hermes, grouped_hermes = process_hermes(df, 'hermes', 'TSH_ES20_HER_Power_Status', column_list, stofile)

    # FIXME
    # - move zero_list and one_list up to here
    # - refactor commonanilty for ER3, ER4, MSG1, and MSG2
    
    ################################
    ### ER3 ###
    # new dataframe (subset) for ER3 (ER3_EE_F04_Power_Status == 'CLOSED')
    df_er3 = dataframe_subset(df, 'er3', 'ER3_EE_F04_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    zero_list = ['off', 'power off', 'opened']
    one_list =  ['on' , 'power on' , 'closed']
    df_er3.ER3_EE_F04_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er3.ER3_EE_F04_Power_Status.values ]        
    
    # pivot to aggregate daily sum for "rack hours" column (similar to grouped_cir)
    grouped_er3 = df_er3.groupby('Date').aggregate(np.sum)
    
    ################################
    ### ER7 ###
    # new dataframe (subset) for ER3 (ER7_EE_F01_Power_Status == 'CLOSED')
    df_er7 = dataframe_subset(df, 'er7', 'ER7_EE_F01_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_er7.ER7_EE_F01_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er7.ER7_EE_F01_Power_Status.values ]        
    
    # pivot to aggregate daily sum for "rack hours" column (similar to grouped_cir)
    grouped_er7 = df_er7.groupby('Date').aggregate(np.sum)    
    
    ################################    
    ### ER5 ###    
    # new dataframe (subset) for ER5 (ER5_Drawer2_Power_Status == 'CLOSED')
    df_er5 = dataframe_subset(df, 'er5', 'ER5_Drawer2_Power_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_er5.ER5_Drawer2_Power_Status = [ normalize_generic(v, one_list, zero_list) for v in df_er5.ER5_Drawer2_Power_Status.values ]    
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_er5 = df_er5.groupby('Date').aggregate(np.sum)

    ################################
    ### MSG OUTLET #1 ###
    # new dataframe (subset) for MSG1 (MSG_Outlet1_Status == 'ON')
    df_msg1 = dataframe_subset(df, 'msg1', 'MSG_Outlet1_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg1.MSG_Outlet1_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg1.MSG_Outlet1_Status.values ]    
    
    # replace NaNs wit zeros
    df_msg1.fillna(0)
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg1 = df_msg1.groupby('Date').aggregate(np.sum)    

    ################################
    ### MSG OUTLET #2 ###
    # new dataframe (subset) for MSG2 (MSG_Outlet2_Status == 'ON')
    df_msg2 = dataframe_subset(df, 'msg2', 'MSG_Outlet2_Status', column_list)
    
    # normalize to change CLOSED to one, and OPENED to zero
    df_msg2.MSG_Outlet2_Status = [ normalize_generic(v, one_list, zero_list) for v in df_msg2.MSG_Outlet2_Status.values ]    
    
    # replace NaNs wit zeros
    df_msg2.fillna(0)
    
    # pivot to aggregate daily sum for "rack hours" column
    grouped_msg2 = df_msg2.groupby('Date').aggregate(np.sum)    

    ################################
    ### CU ###
    # get dataframe with CU hour count info
    cu = CuMonthlyQuery(_HOST_SAMS, _SCHEMA_SAMS, _UNAME_SAMS, _PASSWD_SAMS, date_min, date_max)
    cu_querystr = cu._get_query(date_min, date_max)
    s = StringIO()
    s.write(str(cu))
    s.seek(0) # "rewind" to the beginning of the StringIO object
    df_cu = pd.read_csv(s, sep='\t', parse_dates=True, index_col = [0])

    ################################
    ### PAD ###    
    # get dataframe for padtimes
    df_tmp = pd.read_csv('/misc/yoda/www/plots/batch/padtimes/padtimes.csv', parse_dates=True, index_col = [0])
    df_pad = df_tmp.filter(regex='Date|.*_hours')

    # FIXME verify PAD's been accounted for in CSV up through last day of last month

    # create wall clock dataframe
    df_wall_clock = grouped_er3.copy(deep=True)

    # rename column to Wall_Clock
    df_wall_clock.rename(columns={'ER3_EE_F04_Power_Status':  'Wall_Clock'}, inplace=True)
    df_wall_clock['Wall_Clock'] = 24
    
    # merge (union via how='outer') all dataframes except for df_pad...
    #bamf_df = df_wall_clock.join([grouped_er3, grouped_er4, grouped_cir, grouped_fir, grouped_msg1, grouped_msg2, df_cu, df_pad])
    bamf_df = df_wall_clock   
    # for gr in [grouped_er3, grouped_er5, grouped_er7, grouped_cir, grouped_fir, grouped_msg1, grouped_msg2, df_cu]:
    for gr in [grouped_er3, grouped_er5, grouped_er7, grouped_cir, grouped_fir, grouped_hermes, grouped_msg1, grouped_msg2, df_cu]:
        bamf_df = bamf_df.merge(gr, left_index=True, right_index=True, how='outer')
        
    #...now merge df_pad too, but now get intersection (based on Date index) using how='inner' this time
    bamf_df = bamf_df.merge(df_pad, left_index=True, right_index=True, how='inner')

    # drop unwanted columns
    unwanted_columns = ['mams_ossbtmf_hours', 'iss_radgse_hours', 'mma_0bba_hours', 'mma_0bbb_hours', 'mma_0bbc_hours', 'mma_0bbd_hours']
    for uc in unwanted_columns:
        bamf_df = bamf_df.drop(uc, 1)
        
    # drop unwanted rows (before first day of last month)
    my_today = datetime.date.today()
    my_this_month_day_one = my_today.replace(day=1)
    my_last_month = my_this_month_day_one - datetime.timedelta(days=1)
    first_day = my_last_month.replace(day=1)
    last_day = my_this_month_day_one - datetime.timedelta(days=1)
    bamf_df = bamf_df.loc[first_day:last_day]

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(xlsxfile, engine='xlsxwriter')

    # Create kpi sheet for monthly percentages
    df_config = read_config_template()
    df_config.to_excel(writer, sheet_name='kpi', index=False)
    
    # Create sheet for raw data (via dataframe)
    bamf_df.to_excel(writer, sheet_name='raw', index=True)
    
    # Close the Pandas Excel writer and output the Excel file.
    writer.save()
    print 'wrote raw and kpi template sheets to %s' % xlsxfile

    # Reckon GMT range and overwrite last row with totals; where last row is day 1 of next month
    # and write GMT range into kpi sheet cell B1
    overwrite_last_row_with_totals(xlsxfile, df_config, bamf_df)
    print 'reckoned GMT range\nmodified and saved %s with formulas and totals' % xlsxfile
    

def main(csvfile, resource_csvfile):
    """produce output csv with per-system monthly sensor hour totals"""
    
    # read resource config csv file
    df_cfg = resource_csv2dataframe(resource_csvfile)
    regex_sensor_hours = '.*' + '_hours|.*'.join(df_cfg['Resource']) + '_hours'
    
    #print df_cfg; raise SystemExit
    
    # read input CSV into big pd.DataFrame
    df = csv2dataframe(csvfile)

    # filter to keep only hours columns (gets rid of bytes columns) for each sensor
    # that shows up in df_cfg's Resource column
    ndf = df.filter(regex='Date|Year|Month|Day|' + regex_sensor_hours)
    
    # pivot to aggregate monthly sum for each "sensor_hours" column
    #t = pd.pivot_table(ndf, rows=['Year','Month'], aggfunc=np.sum)
    t = pd.pivot_table(ndf, index=['Year','Month'], aggfunc=np.sum)
    
    # drop the unwanted "Day" column
    df_monthly_hours = t.drop('Day', 1)
    
    # convert index, which are tuples like (YEAR, MONTH), to tuples like (YEAR, MONTH, 1)
    date_tuples = [ ( t + (1,) ) for t in df_monthly_hours.index.values ]

    # convert date_tuples each to hours_in_month
    hours = [ hours_in_month( datetime.date( *tup ) ) for tup in date_tuples ]

    # before we add hours_in_month column, get list of columns for iteration below
    cols = df_monthly_hours.columns.tolist()
    
    # now we can append month_hours column
    df_monthly_hours['hours_in_month'] = pd.Series( hours, index=df_monthly_hours.index)
    
    # iterate over columns (above code helps exclude hours_in_month) to get 100*sensor_hours/hours_in_month
    for c in cols:
        pctstr = c + '_pct'
        pct = 100 * df_monthly_hours[c] / df_monthly_hours['hours_in_month']
        df_monthly_hours[pctstr] = pd.Series( np.round(pct, decimals=0), index=df_monthly_hours.index)
    
    # save csv output file
    csvout = csvfile.replace('.csv', '_monthly_hours.csv')
    df_monthly_hours.to_csv(csvout)
    print 'wrote %s' % csvout


def sto2dataframe(stofile, msid_map):
    """read ascii sto file into dataframe"""
    s = StringIO()
    with open(stofile, 'r') as f:
        # Read and ignore header lines
        header = f.readline() # labels
        s.write(header)
        is_data = False
        for line in f:
            if line.startswith('#Start_Data'):
                is_data = True
            if line.startswith('#End_Data'):
                is_data = False
            if is_data and not line.startswith('#Start_Data'):
                s.write(line)
    s.seek(0) # "rewind" to the beginning of the StringIO object
    df = pd.read_csv(s, sep='\t')
    
    # drop the unwanted "#Header" column
    df = df.drop('#Header', 1)
    column_labels = [ s.replace('Timestamp : Embedded GMT', 'GMT') for s in df.columns.tolist()]
    df.columns = column_labels
    
    # drop Unnamed columns
    for clabel in column_labels:
        if clabel.startswith('Unnamed'):
            df = df.drop(clabel, 1)

    for k, v in msid_map.iteritems():
        df.rename(columns={k: v}, inplace=True)

    return df


if __name__ == '__main__':

    if len(sys.argv) == 2:

        # read input file
        sto_file = sys.argv[1]
        print 'processing ' + sto_file

        # convert to dataframe
        df = sto2dataframe(sto_file, MY_MSID_MAP)

        # normalize power values (on -> 1, off -> 0)...JUST ONE DUMMY COLUMN FOR NOW
        # df.IOP_MP_PWR_CTRL_SAMS = [normalize_cir_fir_power(v) for v in df.IOP_MP_PWR_CTRL_SAMS.values]
        df.CIR_LABS3_Pwr_Hours = [normalize_cir_fir_power(v) for v in df.CIR_LABS3_Pwr_Hours.values]
        df.FIR_LABS4_Pwr_Hours = [normalize_cir_fir_power(v) for v in df.FIR_LABS4_Pwr_Hours.values]

        # drop status columns
        for clabel in df.columns:
            if clabel.startswith('status'):
                df = df.drop(clabel, 1)

        # convert like 2019:335:00:00:10 to datetime objects
        df['dtm'] = [doytimestr_to_datetime(doy_gmtstr) for doy_gmtstr in df.GMT]

        # convert like 2019:335:00:00:10 to date objects
        df['Date'] = [doytimestr_to_datetime(doy_gmtstr).date() for doy_gmtstr in df.GMT]

        # convert datetimes to str and overwrite GMT with those strings
        df['GMT'] = [d.strftime('%Y-%m-%d/%j,%H:%M:%S') for d in df.dtm]

        # drop last row (i.e. drop for for "first day of next month")
        df.drop(df.tail(1).index, inplace=True)

        # create pivot table
        table = pd.pivot_table(df, values=['CIR_LABS3_Pwr_Hours', 'FIR_LABS4_Pwr_Hours'], margins=True, margins_name="Total", index='Date', aggfunc=np.sum)

        # convert tens-of-seconds to hours
        table.CIR_LABS3_Pwr_Hours = table.CIR_LABS3_Pwr_Hours / 360.0
        table.FIR_LABS4_Pwr_Hours = table.FIR_LABS4_Pwr_Hours / 360.0

        # append a totals row
        table.append(table.sum(numeric_only=True), ignore_index=True)

        # convert table to CSV
        table.to_csv(sto_file.replace('.sto', '.csv'))

        # create a Pandas Excel writer using XlsxWriter as the engine.
        xlsx_file = sto_file.replace('.sto', '.xlsx')
        writer = pd.ExcelWriter(xlsx_file, engine='xlsxwriter')

        # # create kpi sheet for monthly percentages
        # df_config = read_config_template()
        # df_config.to_excel(writer, sheet_name='kpi', index=False)

        # create sheet for raw data (via dataframe)
        table.to_excel(writer, sheet_name='raw', index=True)

        # close the Pandas Excel writer and output the Excel file.
        writer.save()
        print 'wrote XLSX file %s' % xlsx_file

        # # pivot to aggregate daily sum for "payload_hours" column
        # column_list = df.columns.tolist()
        # df_cir, grouped_cir = process_cir_fir(df, 'cir', 'CIR_LABS3_Pwr_Hours', column_list, sto_file)
        # df_fir, grouped_fir = process_cir_fir(df, 'fir', 'FIR_LABS4_Pwr_Hours', column_list, sto_file)
        # # df_cir, grouped_cir = process_cir_fir(df, 'fir', 'IOP_MP_PWR_CTRL_SAMS', column_list, sto_file)
        # print df_fir

        # ndf = df.filter(regex='GMT|Date|Year|Month|Day|' + '.*_Pwr')
        # cols = [i for i in ndf.columns if i not in ['GMT', 'Date', 'Year', 'Month', 'Day']]
        # t = pd.pivot_table(ndf, rows=['GMT'], values=cols, aggfunc=np.sum)
        # series = t.transpose().sum()
        # df.to_csv(sto_file.replace('.sto', '.csv'))

        # print df.IOP_MP_PWR_CTRL_SAMS

    else:
        print 'no input (sto) file specified on command line'
        raise SystemExit

    # main(csvfile, resource_csvfile)
