#!/usr/bin/env python

import datetime
from datetime import date
from infinity import inf
from intervals import DateInterval
import pandas as pd
from itertools import islice
from copy import copy

from openpyxl import Workbook, styles
from openpyxl import load_workbook


class MyDateInterval(DateInterval):

    @property
    def radius(self):
        if self.length == inf:
            return inf
        return self.length / 2

    @property
    def third(self):
        if self.length == inf:
            return inf
        return self.length / 3


def demo():

    di = MyDateInterval([date(2018, 1, 1), date(2018, 2, 1)])
    print di.radius

    di2 = MyDateInterval([date(2018, 1, 1), date(2018, 1, 2)])
    print di2.third


def OLDdemo_openpyxl_pandas():
    wb = load_workbook(filename='/Users/ken/Documents/20190204_sample.xlsx')
    ws = wb['Sheet1']
    sheet_ranges = wb['Sheet1']
    print sheet_ranges['F2'].value
    data = ws.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)
    print df
    print df.columns


def update_zin_spreadsheet(filename):
    """update zin spreadsheet"""

    # load workbook and get reference to worksheet
    wb = load_workbook(filename=filename)
    ws = wb['Sheet1']
    # sheet_ranges = wb['Sheet1']

    # TODO verify first 2 rows are identical (an artifact from copying web page)

    # delete row 1
    ws.delete_rows(1, 1)

    # fill cell A1
    ws['A1'] = 'Category'

    # reformat A1 and B1 like C1
    from_to = [('C1', 'A1'), ('C1', 'B1')]
    for c1, c2 in from_to:
        ws[c2].font = copy(ws[c1].font)
        ws[c2].border = copy(ws[c1].border)
        ws[c2].fill = copy(ws[c1].fill)
        ws[c2].number_format = copy(ws[c1].number_format)
        ws[c2].protection = copy(ws[c1].protection)
        ws[c2].alignment = copy(ws[c1].alignment)

    # get dataframe values
    data = ws.values
    cols = next(data)
    data = list(data)
    # idx = [r[0] for r in data]
    data = (islice(r, 0, None) for r in data)
    df = pd.DataFrame(data, columns=cols)
    # print df
    # print df.columns

    # rename Current Investment Mix to Pct and reformat values
    df = df.rename(index=str, columns={"Current Investment Mix": "Current Investment Pct"})
    ws['C1'] = 'Current Investment Pct'

    # verify header row matches
    expected_columns = [u'Category',
                        u'Inv Manager or Sub-Advisor\nInvestment Option',
                        u'Current Investment Pct',
                        u'Units/Shares',
                        u'Unit Value/Share Price',
                        u'Total Balance']
    if all(df.columns == expected_columns):
        print 'Got %d matching column names' % len(expected_columns)
    else:
        print 'hey wait, our column names do not match'

    # get rid of '\nPerformance Snapshot' suffix
    fstr = 'Inv Manager or Sub-Advisor\nInvestment Option'
    df[fstr] = df[fstr].apply(lambda s: s.replace('\nPerformance Snapshot', ''))
    for idx, row in enumerate(df[fstr]):
        cell = 'B%d' % (idx + 2)
        ws[cell] = row

    # ws.column_dimensions['B'].width = 196

    # verify current investment mix totals 100%
    df['Current Investment Pct'] = df['Current Investment Pct'].apply(lambda x: float(x[:-1])/100.0)
    cim = df['Current Investment Pct'].sum()
    if cim == 1.0:
        print 'Current Investment Pct adds up to {:,.2f}%'.format(cim * 100.0)
    else:
        print 'not so fast, our current investment pct does not total to 100%'
    ws['C11'].value = '=sum(C2:C10)'
    ws['C11'].number_format = '#.00'

    # write pct values as floats (not strings)
    for idx, row in enumerate(df['Current Investment Pct']):
        cell = 'C%d' % (idx + 2)
        # print idx, row, ws[cell]
        ws[cell] = row * 100.0
        ws[cell].number_format = '#.00'

    # rewrite dollar amounts as floats (not strings)
    df['Total Balance'] = df['Total Balance'].apply(lambda x: float(x.replace('$', '').replace(',', '')))

    # write dollar values as currency (not strings)
    for idx, row in enumerate(df['Total Balance']):
        cell = 'F%d' % (idx + 2)
        # print idx, float(row.replace('$', '').replace(',', ''))
        ws[cell] = row
        ws[cell].number_format = '$#,##0.00'

    # insert grand total balance at bottom
    total = df['Total Balance'].sum()
    print 'Grand Total Balance is ${:,.2f}'.format(total)
    ws['F11'].value = '=sum(F2:F10)'
    ws['F11'].number_format = '$#,##0.00'

    # add new columns
    ws['G1'] = 'Transfer Out\nof Cash Fund\nDollars'
    ws['H1'] = 'Transfer Out\nof Cash Fund\nPct'
    ws['I1'] = 'Transfer Into\nPct'
    ws['J1'] = 'Transfer Into\nDollars'
    ws['K1'] = 'New\nBalance'
    ws['L1'] = 'New\nPct'
    ws['M1'] = 'Target\nPct'
    ws['H2'].value = 25  # transfer this pct out of cash
    ws['G2'].value, ws['G2'].number_format = '=F2*H2/100', '$#,##0.00'
    ws['K2'].value = '=F2-G2'
    ws['I3'].value = 30
    ws['I4'].value = 20
    ws['I5'].value = 0
    ws['I6'].value = 30
    ws['I7'].value = 0
    ws['I8'].value = 5
    ws['I9'].value = 15
    ws['I10'].value = 0
    ws['K2'].value, ws['K2'].number_format = '=F2-G2', '$#,##0.00'

    for r in range(2, 11):
        c1 = 'L%d' % r
        ws[c1].value = '=100*K%d/$F$11' % r
        ws[c1].number_format = '#.00'

    for r in range(3, 11):
        c1, c2 = 'J%d' % r, 'K%d' % r
        frm1, frm2 = '=$G$2*I%d/100' % r, '=F%d+J%d' % (r, r)
        ws[c1].value, ws[c2].value = frm1, frm2
        ws[c1].number_format, ws[c2].number_format = '$#,##0.00', '$#,##0.00'
    ws['I11'].value = '=SUM(I3:I10)'
    ws['J11'].value, ws['J11'].number_format = '=SUM(J3:J10)', '$#,##0.00'
    ws['K11'].value, ws['K11'].number_format = '=SUM(K2:K10)', '$#,##0.00'
    ws['M11'].value, ws['M11'].number_format = '=SUM(M2:M10)', '##0'

    # create alignment styles
    wrap_alignment = styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
    right_alignment = styles.Alignment(horizontal='right')

    ws['H3'].value = 'adjust pct in cell above'
    ws['I2'].value = 'adjust pct in cells below'
    ws['H3'].alignment = wrap_alignment
    ws['I2'].alignment = wrap_alignment
    ws['H3'].font = styles.Font(color=styles.colors.WHITE)
    ws['I2'].font = styles.Font(color=styles.colors.WHITE)

    cidx = '00808080'  # styles.colors.COLOR_INDEX[2]
    my_color = styles.colors.Color(rgb=cidx)
    my_fill = styles.fills.PatternFill(patternType='solid', fgColor=my_color)
    ws['I2'].fill = my_fill
    ws['J2'].fill = my_fill

    for r in range(3, 11):
        c1, c2 = 'G%d' % r, 'H%d' % r
        ws[c1].fill, ws[c2].fill = my_fill, my_fill

    ws['M2'].value = 5
    ws['M3'].value = 15
    ws['M4'].value = 25
    ws['M5'].value = 20
    ws['M6'].value = 10
    ws['M7'].value = 5
    ws['M8'].value = 5
    ws['M9'].value = 10
    ws['M10'].value = 5

    # TODO add the following cross-checks
    # I12               J12             K12
    # I11 == 100?       J11 == G2?      K11 == F11?

    ws['C12'].value = '=if(C11 = 100, "yes", "no")'
    ws['C12'].alignment = right_alignment

    ws['I12'].value = '=if(I11 = 100, "yes", "no")'
    ws['I12'].alignment = right_alignment

    ws['J12'].value = '=if(J11 = G2, "yes", "no")'
    ws['J12'].alignment = right_alignment

    ws['K12'].value = '=if(K11 = F11, "yes", "no")'
    ws['K12'].alignment = right_alignment

    ws['M12'].value = '=if(M11 = 100, "yes", "no")'
    ws['M12'].alignment = right_alignment

    # Show some color cells with COLOR_INDEX values too
    # rr = 2
    # for cidx in styles.colors.COLOR_INDEX:
    #     my_color = styles.colors.Color(rgb=cidx)
    #     my_fill = styles.fills.PatternFill(patternType='solid', fgColor=my_color)
    #     ws['M%d' % rr].fill, ws['N%d' % rr].value = my_fill, cidx
    #     rr += 1
    #     if rr > 63:
    #         break

    # copy format to newly inserted cim & grand total balance
    from_to = [('A10', 'A11'),
               ('B10', 'B11'),
               ('F10', 'F11'),
               ('C10', 'C11'),
               ('F1',  'G1'),
               ('F1',  'H1'),
               ('F1',  'I1'),
               ('F1',  'J1'),
               ('F1',  'K1'),
               ('F1',  'L1'),
               ('F1',  'M1')]
    for c1, c2 in from_to:
        ws[c2].font = copy(ws[c1].font)
        ws[c2].border = copy(ws[c1].border)
        ws[c2].fill = copy(ws[c1].fill)
        ws[c2].number_format = copy(ws[c1].number_format)
        ws[c2].protection = copy(ws[c1].protection)
        ws[c2].alignment = copy(ws[c1].alignment)

    ws['A11'] = 'Combined'
    ws['B11'] = 'All Funds'
    rd = ws.row_dimensions[11]  # get dimension for row 11
    rd.height = 25  # value in points, there is no "auto"

    # adjust row heights
    ws.row_dimensions[1].height = 48
    for inr in range(2, 11):
        ws.row_dimensions[inr].height = 56

    # save new file
    out_file = filename.replace('_raw.xlsx', '_change.xlsx')
    wb.save(out_file)


def demo_openpyxl():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("/Users/ken/Documents/xxx20190207_sample_out.xlsx")


if __name__ == '__main__':
    # import pyperclip
    # hoping = pyperclip.paste()
    # print hoping
    # raise SystemExit

    update_zin_spreadsheet('/Users/ken/Documents/2019-02-16_zin_raw.xlsx')

    # demo()
