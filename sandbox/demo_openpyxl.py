#!/usr/bin/env python

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
from openpyxl.reader.excel import load_workbook
import openpyxl.style

def demo_format():
    
    wb = Workbook()
    dest_filename = r'/tmp/1empty_book.xlsx'
    ws = wb.worksheets[0]    
    
    name = 'name'
    
    i = 0
    for col_idx in xrange(1, 9):
        col = get_column_letter(col_idx)
        for row in xrange(1, 11):
            ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)
            ws.cell('%s%s'%(col, row)).style.fill.fill_type = 'solid'
            ws.cell('%s%s'%(col, row)).style.fill.start_color.index = openpyxl.style.Color.DARKYELLOW
            i += 1
            wb.create_named_range(name+str(i), ws, '%s%s'%(col, row))
    
    wb.save(filename = dest_filename)    

demo_format(); raise SystemExit

def demo_open_existing():
    wb = load_workbook(filename = r'/tmp/empty_book.xlsx')
    ws = wb.get_sheet_by_name("ER3raw")
    #ws.cell(row=0, column=0).delete()
    print ws.cell('A19').value
    #ws.cell('A19').value = '/</-/'
    #wb.save(filename = r'/tmp/empty_book.xlsx')

def demo_create_write():
    wb = Workbook()
    
    dest_filename = r'/tmp/empty_book.xlsx'
    
    ws = wb.worksheets[0]
    
    ws.title = "range names"
    
    for col_idx in xrange(1, 40):
        col = get_column_letter(col_idx)
        for row in xrange(1, 600):
            ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)
    
    ws = wb.create_sheet()
    
    ws.title = 'Pi'
    
    ws.cell('F5').value = 3.14
    
    wb.save(filename = dest_filename)
        
def update_xlsx(src, dest):
    #Open an xlsx for reading
    wb = load_workbook(filename = src)
    
    ##Get the current Active Sheet
    #ws = wb.get_active_sheet()
    
    #You can also select a particular sheet name
    ws = wb.get_sheet_by_name("ER3grp")
    ws.cell(row=0, column=0).delete()
    
    ##Open the csv file
    #with open(src) as fin:
    #    #read the csv
    #    reader = csv.reader(fin)
    #    #enumerate the rows, so that you can
    #    #get the row index for the xlsx
    #    for index,row in enumerate(reader):
    #        #Asssuming space sepeated,
    #        #Split the row to cells (column)
    #        row = row[0].split()
    #        #Access the particular cell and assign
    #        #the value from the csv row
    #        ws.cell(row=index,column=7).value = row[2]
    #        ws.cell(row=index,column=8).value = row[3]
            
    #save the file
    wb.save(dest)
    
if __name__ == "__main__":
    #update_xlsx('/tmp/empty_book.xlsx', '/tmp/new_empty.xlsx')
    demo_open_existing()